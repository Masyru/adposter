from openpyxl import load_workbook
from json import dump

# Parse the xlsx files from japancar with models and firm

def parse_first():
    # Load in the workbook
    wb = load_workbook('./1.xlsx')

    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    rest = {}

    for i in range(1, 4195):
        name = sheet.cell(row=i, column=1).value
        print(i)
        if name in rest.keys():
            rest[name].append(str(sheet.cell(row=i, column=2).value))
        else:
            rest[name] = [str(sheet.cell(row=i, column=2).value)]

    with open('./models.json', 'w') as fout:
        dump(rest, fout)


def parse_second():
    wb = load_workbook('./2.xlsx')

    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    rest = []

    for i in range(1, 2070):
        rest.append(sheet.cell(row=i, column=1).value)
        print(i)

    with open('./parts.json', 'w') as fout:
        dump(rest, fout)

